import openpyxl as op


def add_password():
    # window os
    path = r"C:\Users\User\Documents\password-handler\file"
    # mac os
    # path = "/Users/yoosunu/Desktop/password-handler/새 폴더"
    wb = op.load_workbook(path + "\password.xlsx")
    # wb = op.load_workbook(path+"/password.csv")
    ws = wb.active

    """Define Infos"""
    datasets = []

    print("사이트의 이름을 입력해주세요: ")
    SITE = input()
    print("아이디를 입력해주세요: ")
    ID = input()
    print("비밀번호를 입력해주세요: ")
    PASSWORD = input()

    column_max = ws.max_column
    """data define"""

    # SITE = "google"
    # ID = "id"
    # PASSWORD = "password"

    datasets.append(SITE)
    datasets.append(ID)
    datasets.append(PASSWORD)

    """data input"""

    if column_max == 1:
        for i, dataset in enumerate(datasets, start=1):
            ws.cell(row=i, column=2, value=dataset)
    elif column_max >= 2:
        for i, dataset in enumerate(datasets, start=1):
            ws.cell(row=i, column=column_max + 1, value=dataset)

    """saving"""
    wb.save(path + "\password.xlsx")
    wb.close()
